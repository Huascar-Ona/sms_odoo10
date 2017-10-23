# -*- coding: utf-8 -*-
from odoo import api,fields,models
import base64
from openpyxl import load_workbook
import os
import tempfile
import magic

class quick_send_wizard(models.TransientModel):
    _name = "sms.quick_send_wizard"
    _description = "Envio rapido"

    name = fields.Char(u"Identificador de envío")
    text = fields.Text("Texto")
    dests = fields.Text("Destinatarios", help="Separar por saltos de línea")
    dbfile = fields.Binary("Archivo destinatarios", help="Puede subir un archivo con los destinatarios")
    schedule_date = fields.Datetime("Fecha programada", help="Cuándo se deben empezar a enviar estos mensajes. Dejar en blanco para enviar en este momento")
    list_id = fields.Many2one('sms.list', string='Nombre de Lista')

    @api.multi
    def action_process(self):
        Sms = self.env["sms.sms"]

        data = {}
        dests = []
        if self.dests:
            for number in self.dests.split("\n"):
                number = number.strip()
                dests.append(number)
        if self.dbfile:
            content = base64.b64decode(self.dbfile)

            handle, path = tempfile.mkstemp(suffix=".xlsx")
            with open(path, "w") as f:
                f.write(content)

            numeros = []

            try:
                wb = load_workbook(filename=path)
            except:
                raise Exception("El archivo enviado no es un archivo Excel válido o compatible")
            first_sheet = wb.get_sheet_names()[0]
            worksheet = wb.get_sheet_by_name(first_sheet)
            irow = 1
            for row in worksheet.iter_rows():
                irow += 1
                for i,cell in enumerate(row):
                    if i == 0:
                        numero = str(cell.value)
                        numeros.append(numero)
                        data[numero] = []
                    else:
                        data[numeros[-1]].append(cell.value)

            print("Numeros: {} \n Data: {}").format(numeros, data)

            dests = numeros

        def format_msg(data, numero, msg):
          if numero in data:
              row = data[numero]
              for datum in row:
                  if type(datum) not in (str,unicode):
                      datum = str(datum)

                  msg = msg.replace(u"XXX", datum, 1)

                  if not u"XXX" in msg:
                      print('No hay XXX')
                      break

              return msg
          else:
              return self.text

        if self.list_id:
            for x in self.env["sms.contact"].search([('list_id','=',self.list_id.id)]):
                dests.append(x.telefono)

        created = []
        if not data:
            data = {numero:[] for numero in dests}
        for number in dests:
            if number:
                created.append(Sms.create({
                    'text': format_msg(data, number, self.text),
                    'dest': number,
                    'schedule_date': self.schedule_date,
                    'name': self.name
                }))

        return {
            'name': 'SMS',
            'res_model': 'sms.sms',
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'list,form',
            'domain': [('id', 'in', [x.id for x in created])]
        }
